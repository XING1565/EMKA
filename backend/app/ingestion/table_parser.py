from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import Path
from typing import Iterable


def parse_table_file(filename: str, content: bytes, metadata: dict) -> str:
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".csv":
        return _parse_csv(content, metadata)
    if suffix == ".xlsx":
        return _parse_xlsx(filename, content, metadata)
    return _parse_csv(content, metadata)


def _parse_csv(content: bytes, metadata: dict) -> str:
    text = content.decode("utf-8-sig", errors="replace")
    rows = list(csv.reader(StringIO(text)))
    return _table_to_text("csv", rows, metadata)


def _parse_xlsx(filename: str, content: bytes, metadata: dict) -> str:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        metadata["sheet_names"] = list(workbook.sheetnames)
        sections: list[str] = []
        total_rows = 0
        row_ranges: list[str] = []
        all_columns: list[str] = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = [[_cell_to_text(value) for value in row] for row in sheet.iter_rows(values_only=True)]
            sheet_metadata: dict = {}
            section = _table_to_text(sheet_name, rows, sheet_metadata)
            if section:
                sections.append(section)
            total_rows += int(sheet_metadata.get("row_count", 0))
            row_ranges.extend(sheet_metadata.get("row_ranges", []))
            for col in sheet_metadata.get("table_columns", []):
                if col not in all_columns:
                    all_columns.append(col)
        metadata["row_count"] = total_rows
        metadata["row_ranges"] = row_ranges
        metadata["table_columns"] = all_columns
        return "\n\n".join(sections) if sections else f"No table rows found in {filename}."
    except Exception as exc:
        metadata["parse_error"] = str(exc)
        return f"XLSX table extraction failed for {filename}: {exc}"


def _table_to_text(label: str, rows: Iterable[list[str]], metadata: dict) -> str:
    cleaned = [[str(cell).strip() for cell in row] for row in rows if any(str(cell).strip() for cell in row)]
    if not cleaned:
        metadata["table_columns"] = []
        metadata["row_count"] = 0
        metadata["row_ranges"] = []
        return ""

    header = cleaned[0]
    data_rows = cleaned[1:]
    metadata["table_columns"] = header
    metadata["row_count"] = len(data_rows)
    metadata["row_ranges"] = [f"2-{len(cleaned)}"] if data_rows else []

    lines = [
        f"Table summary ({label}): {len(data_rows)} data rows, columns: {', '.join(header)}",
    ]
    for row_number, row in enumerate(data_rows, start=2):
        pairs = []
        for index, value in enumerate(row):
            column = header[index] if index < len(header) and header[index] else f"column_{index + 1}"
            pairs.append(f"{column}={value}")
        lines.append(f"row {row_number}: " + "; ".join(pairs))
    return "\n".join(lines)


def _cell_to_text(value) -> str:
    if value is None:
        return ""
    return str(value)
